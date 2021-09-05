from Data_Plotter import Excel
from openpyxl import load_workbook
from datetime import date,datetime

import os

### This Package works on setting up the program
# and make new worksheet at begining of the month and delete
# old sheets which are 12 months old###

XL = Excel()
today = date.today()

class First_Setup:

    def setup(self):
        Workbook_status = os.path.isfile(XL.filename)
        if not Workbook_status:
            XL.setup_workbook()

    def previous_data(self, previous_amount):
        month = today.month
        p_month = month - 1
        if p_month < 1:
            p_month = 12

        datetime_object = datetime.strptime(p_month, "%m")
        prev_month = datetime_object.strftime("%B")
        date = today.strftime("%m/%d")
        XL.add_data('Yearly', prev_month, previous_amount)
        XL.new_sheet(today.strftime("%B"), today.strftime("%Y") )

class Regular_Check:

    def check(self):
        wb = load_workbook(XL.filename)
        existing = wb.sheetnames
        current_month = today.strftime("%B")

        if current_month not in existing:
            month = today.month
            p_month = month - 1
            if p_month < 1:
                p_month = 12

            datetime_object = datetime.strptime(p_month, "%m")
            prev_month = datetime_object.strftime("%B")
            pmws = wb[prev_month]
            prev_month_total = pmws['C3'].value
            title = pmws['A1'].value
            yr_particular = title.replace("Expenses for ","")
            # c_date = today.strftime("%m/%d")
            XL.add_data('Yearly', yr_particular, prev_month_total)

            XL.new_sheet(current_month, today.strftime("%Y"))

            d_month = month - 2
            if d_month < 1:
                d_month = 12 - d_month

            del_month = calendr.moanth_name[d_month]
            if del_month in existing:
                XL.delete_sheet(del_month)






