from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import date

today = date.today()

class Excel:
    filename = 'test2.xlsx'


    def setup_workbook(self):
        """Used at initialization of program to make a new workbook when used first time"""
        constant_flag = False
        wb = Workbook()
        existing = wb.sheetnames
        for names in existing:
            if names == 'Recurring Constant':
                constant_flag = True

            if names == 'Sheet':
                del wb['Sheet']

        if not constant_flag:
            required_sheets = ['Recurring Constant', 'Yearly']
            for make_sheet in required_sheets:
                wb.create_sheet(make_sheet)
                cs = wb[make_sheet]
                cs.merge_cells('A1:D1')

                cs['A1'] = make_sheet + " Expenses"
                cs['A1'].font = Font(size='20', bold=True)

                cs['B3'] = 'Total '+ make_sheet +' Amount'
                cs['B3'].font = Font(bold=True)

                cs['C3'] = 0
                cs['C3'].font = Font(bold=True)

                cs['A6'] = 'Updated Date'
                cs['B6'] = 'Particulars'
                cs['C6'] = 'Amount'
                cs['B6'].font = Font(bold=True)
                cs['C6'].font = Font(bold=True)
                cs['A6'].font = Font(bold=True)

                cs.column_dimensions['A'].width = 20
                cs.column_dimensions['B'].width = 20
                cs.column_dimensions['B'].width = 15

                cs['A6'].fill = PatternFill('solid', fgColor="B8DFD8")
                cs['B6'].fill = PatternFill('solid', fgColor="B8DFD8")
                cs['C6'].fill = PatternFill('solid', fgColor="B8DFD8")

            wb.save(self.filename)

    def new_sheet(self,month, year):
        """Creates a new sheet, plots all the desired templates and feeds the previous month total"""
        wb = load_workbook(self.filename)
        wb.create_sheet(month)

        cs = wb[month]
        cs.merge_cells('A1:E1')

        cs['A1'] = "Expenses for " + month + " " +year
        cs['A1'].font = Font(size='20', bold=True)

        cs['B3'] = 'Total Expenses'
        cs['B3'].font = Font(bold=True)

        cs['C3'] = 0
        cs['C3'].font = Font(bold=True)

        cs['D3'] = 0
        cs['D3'].font = Font(bold=True)

        #For feeding the previous month total
        cs['B4'] = 'Previous Month'
        cs['B4'].font = Font(bold=True)

        previous_month_total = self.get_previous()

        cs['C4'] = previous_month_total
        cs['C4'].font = Font(bold=True)

        cs['A6'] = 'DATE'
        cs['B6'] = 'PARTICULARS'
        cs['C6'] = 'AMOUNT'
        cs['D6'] = 'WITHDRAWN'

        cs['B6'].font = Font(bold=True)
        cs['C6'].font = Font(bold=True)
        cs['A6'].font = Font(bold=True)
        cs['D6'].font = Font(bold=True)

        cs.column_dimensions['A'].width = 15
        cs.column_dimensions['B'].width = 20
        cs.column_dimensions['c'].width = 20
        cs.column_dimensions['D'].width = 20

        cs['A6'].fill = PatternFill('solid', fgColor="B8DFD8")
        cs['B6'].fill = PatternFill('solid', fgColor="B8DFD8")
        cs['C6'].fill = PatternFill('solid', fgColor="B8DFD8")
        cs['D6'].fill = PatternFill('solid', fgColor="B8DFD8")

        wb.save(self.filename)

    def add_data(self,base_sheet, particular = None, amount = None, withdrawn=None):
        """Adds data into a worksheet"""
        wb = load_workbook(self.filename)
        ws = wb[base_sheet]
        date = today.strftime('%d %b')

        if amount != None:
            amount = int(amount)
            data = (date, particular, amount, withdrawn)
            ws.append(data)

            total = ws['C3'].value
            total = int(total) + amount
            ws['C3'] = int(total)



        if withdrawn != None:
            withdrawn = int(withdrawn)
            data = (date, particular, amount, withdrawn)
            ws.append(data)

            total = ws['C4'].value
            total = int(total) + withdrawn
            ws['C3'] = int(total)



        wb.save(self.filename)

    def delete_sheet(self,sheet_name):
        """Deletes a sheet from the workbook"""
        wb = load_workbook(self.filename)
        del wb[sheet_name]
        wb.save(self.filename)

    def get_previous(self):
        wb = load_workbook(self.filename)
        ws = wb['Yearly']
        last = ws.max_row
        index = 'C'+ str(last)
        data = ws[index].value
        return data

    def get_paricular_list(self,base_sheet):
        """Returns a list of all the particulars in a sheet"""
        wb = load_workbook(self.filename)
        ws = wb[base_sheet]
        rows = ws.max_row
        data_list = []

        for x in range(7,rows+1):
            index = 'B'+ str(x)
            data = ws[index].value
            data_list.append(data)

        return data_list

    def get_amount_list(self,base_sheet):
        """Returns a list of all the amount in a sheet"""
        wb = load_workbook(self.filename)
        ws = wb[base_sheet]
        rows = ws.max_row
        data_list = []

        for x in range(7,rows+1):
            index = 'C'+ str(x)
            data = ws[index].value
            data_list.append(data)

        return data_list

    def read_previous(self,month):
        wb = load_workbook(self.filename)
        ws = wb[month]
        data = ws['C4'].value
        return data

    def read_total(self,month):
        wb = load_workbook(self.filename)
        ws = wb[month]
        data = ws['C3'].value
        return int(data)

    def get_value(self, base_sheet, index):
        wb = load_workbook(self.filename)
        ws = wb[base_sheet]
        data = ws[index].value
        return data

    def update_data(self, index, value):
        wb = load_workbook(self.filename)
        ws = wb['Recurring Constant']
        ws[index] = int(value)
        wb.save(self.filename)

    def remove_data(self,base_sheet, row_id):
        wb = load_workbook(self.filename)
        ws = wb[base_sheet]
        ws.delete_rows(row_id)
        wb.save(self.filename)

    def percent_saving(self, this_month, prev_month):
        percentage = (prev_month - this_month)/prev_month * 100

        if percentage >= 0:
            return 'Saved %.2f' %percentage +'%'

        if percentage < 0:
            percentage = -1 * percentage
            return 'Spent %.2f'%percentage + '% more'

    def get_sheets(self):
        wb = load_workbook(self.filename)
        return wb.sheetnames











